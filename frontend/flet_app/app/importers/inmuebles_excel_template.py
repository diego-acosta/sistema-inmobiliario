from __future__ import annotations

from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_SHEET: Final = "Datos"
HELP_SHEET: Final = "Ayuda"

EXPECTED_HEADERS: Final[tuple[str, ...]] = (
    "codigo",
    "descripcion",
    "desarrollo",
    "calle",
    "altura",
    "estado_administrativo",
    "estado_juridico",
    "m2",
    "observaciones",
    "folio_real",
    "circunscripcion",
    "seccion",
    "chacra",
    "quinta",
    "fraccion",
    "manzana",
    "lote",
    "parcela",
    "subparcela",
    "nomenclatura_catastral",
    "nomenclatura_madre",
    "partida",
    "matricula",
    "superficie_titulo",
    "superficie_mensura",
    "medidas",
    "situacion_posesoria",
    "situacion_dominial",
    "organismo_origen",
    "fecha_desde",
    "fecha_hasta",
    "estado_dato",
    "observaciones_catastrales",
)

EXAMPLE_ROWS: Final[tuple[tuple[str, ...], ...]] = (
    (
        "IMP-001", "Lote importado básico", "Loteo existente", "San Martín", "123", "ACTIVO", "REGULAR", "500", "Fila básica sin dato catastral avanzado",
        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
    ),
    (
        "IMP-002", "Lote con catastro avanzado", "Loteo existente", "Ruta 8", "S/N", "ACTIVO", "REGULAR", "1234.56", "Fila con dato catastral/registral completo",
        "FR-001", "I", "A", "CH-1", "Q-2", "F-3", "M1", "L2", "P-10", "SP-1", "NC-001", "NC-MADRE-001", "P-001", "MAT-001",
        "1200.50", "1198.75", "20x60", "POSESION REGULAR", "DOMINIO PERFECTO", "Catastro municipal", "2026-01-01", "", "ACTIVO", "Observación catastral ejemplo",
    ),
    (
        "IMP-003", "Ejemplo con estructura completa y vacíos", "", "", "", "", "", "750", "",
        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
    ),
)

COLUMN_HELP: Final[tuple[tuple[str, str, str, str, str], ...]] = (
    ("codigo", "Obligatoria", "Texto", "Identificador del inmueble. No debe repetirse ni existir previamente.", "codigo, código, codigo_lote, cod_lote, lote_codigo"),
    ("descripcion", "Opcional", "Texto", "Nombre o descripción visible del inmueble/lote.", "descripcion, descripción, nombre, nombre_lote"),
    ("desarrollo", "Opcional", "Texto", "Código o nombre de un desarrollo existente si se informa.", "desarrollo, loteo, emprendimiento, barrio"),
    ("calle", "Opcional", "Texto", "Calle del inmueble para dirección básica estructurada.", "calle, nombre_calle, via"),
    ("altura", "Opcional", "Texto", "Altura del inmueble; admite valores como S/N o 123 bis.", "altura, numero, nro, número"),
    ("estado_administrativo", "Opcional", "Texto", "Estado administrativo del inmueble; si falta se usa ACTIVO.", "estado_administrativo, estado, estado_admin"),
    ("estado_juridico", "Opcional", "Texto", "Estado jurídico del inmueble; si falta se usa REGULAR.", "estado_juridico, estado legal, situacion juridica, situación jurídica"),
    ("m2", "Opcional", "Decimal positivo", "Superficie del inmueble.", "superficie, superficie_m2, m2, metros"),
    ("observaciones", "Opcional", "Texto", "Observaciones del inmueble básico.", "observaciones, obs, comentarios"),
    ("folio_real", "Opcional", "Texto", "Folio real del dato registral.", "folio_real"),
    ("circunscripcion", "Opcional", "Texto", "Circunscripción catastral.", "circunscripcion, circunscripción, circ"),
    ("seccion", "Opcional", "Texto", "Sección catastral.", "seccion, sección, sec"),
    ("chacra", "Opcional", "Texto", "Chacra catastral.", "chacra"),
    ("quinta", "Opcional", "Texto", "Quinta catastral.", "quinta"),
    ("fraccion", "Opcional", "Texto", "Fracción catastral.", "fraccion, fracción"),
    ("manzana", "Opcional", "Texto", "Manzana usada para crear dato catastral/registral.", "manzana, mz, mza"),
    ("lote", "Opcional", "Texto", "Dato funcional/catastral del inmueble; no crea una entidad lote separada.", "lote, nro_lote, numero_lote, número_lote"),
    ("parcela", "Opcional", "Texto", "Parcela catastral.", "parcela"),
    ("subparcela", "Opcional", "Texto", "Subparcela catastral.", "subparcela, sub_parcela"),
    ("nomenclatura_catastral", "Opcional", "Texto", "Nomenclatura catastral.", "nomenclatura_catastral, nomenclatura, nomenclatura catastro"),
    ("nomenclatura_madre", "Opcional", "Texto", "Nomenclatura madre u origen del inmueble/lote, si deriva de una nomenclatura anterior.", "nomenclatura_madre, nomenclatura madre, nom_madre, nomenclatura origen, nomenclatura_origen"),
    ("partida", "Opcional", "Texto", "Partida inmobiliaria.", "partida, partida_inmobiliaria"),
    ("matricula", "Opcional", "Texto", "Matrícula registral.", "matricula, matrícula, matricula_registral"),
    ("superficie_titulo", "Opcional", "Decimal positivo", "Superficie según título.", "superficie_titulo, superficie título, superficie_de_titulo"),
    ("superficie_mensura", "Opcional", "Decimal positivo", "Superficie según mensura.", "superficie_mensura, superficie_de_mensura"),
    ("medidas", "Opcional", "Texto", "Medidas o dimensiones descriptivas.", "medidas"),
    ("situacion_posesoria", "Opcional", "Texto", "Situación posesoria.", "situacion_posesoria, situación posesoria"),
    ("situacion_dominial", "Opcional", "Texto", "Situación dominial.", "situacion_dominial, situación dominial"),
    ("organismo_origen", "Opcional", "Texto", "Organismo de origen del dato catastral/registral.", "organismo_origen, organismo"),
    ("fecha_desde", "Opcional", "Fecha AAAA-MM-DD", "Fecha informativa del dato catastral/registral principal; no abre historial formal ni se completa automáticamente.", "fecha_desde, vigencia_desde"),
    ("fecha_hasta", "Opcional", "Fecha AAAA-MM-DD", "Fecha informativa del dato catastral/registral principal; no cierra vigencias automáticamente y no puede ser anterior a fecha_desde.", "fecha_hasta, vigencia_hasta"),
    ("estado_dato", "Opcional", "ACTIVO/INACTIVO/HISTORICO", "Estado informativo del dato principal; si falta se usa ACTIVO y no dispara historial formal.", "estado_dato, estado_dato_catastral"),
    ("observaciones_catastrales", "Opcional", "Texto", "Observaciones propias del dato catastral/registral.", "observaciones_catastrales, observaciones_catastro, observaciones_registrales"),
)

VALIDATION_RULES: Final[tuple[str, ...]] = (
    "codigo es obligatorio.",
    "codigo no debe repetirse en el archivo.",
    "codigo no debe existir previamente en el sistema.",
    "m2, superficie_titulo y superficie_mensura deben ser números positivos si se informan.",
    "fecha_desde y fecha_hasta son opcionales e informativas para el dato catastral/registral principal; no implementan historial formal.",
    "Si se informan, fecha_desde y fecha_hasta deben ser fechas válidas; se recomienda AAAA-MM-DD y también se aceptan DD/MM/YYYY o DD-MM-YYYY.",
    "fecha_hasta no puede ser anterior a fecha_desde.",
    "desarrollo debe coincidir con código o nombre de un desarrollo existente.",
    "lote no es una entidad separada; se trata como dato funcional/catastral del inmueble.",
    "Las columnas catastrales/registrales vacías se omiten del payload.",
    "No se gestionan vigencias múltiples ni cierre automático de datos catastrales/registrales.",
    "No se importan ventas.",
    "No se importan precios.",
    "No se importan servicios.",
    "No se importa geometría/plano.",
)


def create_inmuebles_excel_template(path: str) -> str:
    """Create the inmuebles/lotes import template and return its absolute path."""
    output = Path(path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    data = workbook.active
    data.title = DATA_SHEET
    _build_data_sheet(data)
    help_sheet = workbook.create_sheet(HELP_SHEET)
    _build_help_sheet(help_sheet)
    workbook.save(output)
    return str(output)


def _build_data_sheet(sheet) -> None:
    sheet.append(EXPECTED_HEADERS)
    for row in EXAMPLE_ROWS:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    _auto_width(sheet)


def _build_help_sheet(sheet) -> None:
    sheet.append(("Plantilla completa de importación de inmuebles/lotes",))
    sheet["A1"].font = Font(size=14, bold=True)
    sheet.append(("",))
    sheet.append(("Reglas de validación",))
    sheet["A3"].font = Font(bold=True)
    for rule in VALIDATION_RULES:
        sheet.append((rule,))
    sheet.append(("",))
    sheet.append(("Columnas", "Obligatoriedad", "Formato esperado", "Descripción", "Aliases aceptados"))
    header_row = sheet.max_row
    for item in COLUMN_HELP:
        sheet.append(item)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    sheet.freeze_panes = f"A{header_row + 1}"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _auto_width(sheet, max_width=72)


def _auto_width(sheet, *, max_width: int = 36) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[letter].width = min(max(width, 12), max_width)
