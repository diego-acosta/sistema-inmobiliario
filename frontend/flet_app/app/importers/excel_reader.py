from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from .excel_import_models import ExcelColumn, ExcelSheetData, ExcelWorkbookData


class ExcelImportError(ValueError):
    """User-safe error raised while reading Excel files."""


def normalize_column_name(text: object) -> str:
    value = "" if text is None else str(text)
    value = unicodedata.normalize("NFKD", value.strip().lower())
    value = "".join(char for char in value if not unicodedata.combining(char))
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r"[^a-z0-9_]+", "", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value


def read_excel_workbook(path: str | Path, sheet_name: str | None = None) -> ExcelWorkbookData:
    file_path = Path(path)
    if file_path.suffix.lower() != ".xlsx":
        raise ExcelImportError("Extensión no soportada. Seleccioná un archivo .xlsx.")
    if not file_path.exists():
        raise ExcelImportError("Archivo no encontrado.")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExcelImportError("No está instalada la dependencia openpyxl.") from exc

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - keep traceback away from UI
        raise ExcelImportError("No se pudo leer el archivo Excel.") from exc

    sheet_names = list(workbook.sheetnames)
    if not sheet_names:
        raise ExcelImportError("El archivo Excel no contiene hojas.")
    active_sheet = sheet_name or workbook.active.title or sheet_names[0]
    if active_sheet not in sheet_names:
        raise ExcelImportError(f"La hoja '{active_sheet}' no existe en el archivo.")

    try:
        selected_sheet = _read_sheet(workbook[active_sheet])
    finally:
        workbook.close()

    return ExcelWorkbookData(
        path=str(file_path),
        sheet_names=sheet_names,
        active_sheet=active_sheet,
        sheets={active_sheet: selected_sheet},
    )


def _read_sheet(sheet: Any) -> ExcelSheetData:
    header_values: list[Any] | None = None
    header_row_number = 0
    data_rows: list[tuple[int, tuple[Any, ...]]] = []

    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_values = tuple(values or ())
        if header_values is None:
            if _row_has_value(row_values):
                header_values = list(row_values)
                header_row_number = row_number
            continue
        data_rows.append((row_number, row_values))

    if header_values is None:
        raise ExcelImportError(f"La hoja '{sheet.title}' no tiene encabezados.")

    columns: list[ExcelColumn] = []
    seen: dict[str, int] = {}
    for index, raw_name in enumerate(header_values):
        if raw_name is None or str(raw_name).strip() == "":
            continue
        name = str(raw_name).strip()
        normalized = normalize_column_name(name)
        if not normalized:
            continue
        seen[normalized] = seen.get(normalized, 0) + 1
        if seen[normalized] > 1:
            normalized = f"{normalized}_{seen[normalized]}"
        columns.append(ExcelColumn(index=index, name=name, normalized_name=normalized))

    if not columns:
        raise ExcelImportError(f"La hoja '{sheet.title}' no tiene encabezados válidos.")

    rows: list[dict[str, Any]] = []
    for row_number, row_values in data_rows:
        if not _row_has_value(row_values):
            continue
        row = {"__row_number__": row_number}
        for column in columns:
            row[column.normalized_name] = row_values[column.index] if column.index < len(row_values) else None
        rows.append(row)

    if not rows:
        raise ExcelImportError(f"La hoja '{sheet.title}' no tiene filas de datos.")

    return ExcelSheetData(
        name=sheet.title,
        columns=columns,
        rows=rows,
        header_row_number=header_row_number,
    )


def _row_has_value(values: tuple[Any, ...]) -> bool:
    return any(value is not None and str(value).strip() != "" for value in values)
