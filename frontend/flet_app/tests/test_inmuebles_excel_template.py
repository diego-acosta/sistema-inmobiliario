from __future__ import annotations

from openpyxl import load_workbook

from app.importers.inmuebles_excel_template import EXPECTED_HEADERS, create_inmuebles_excel_template


def test_crea_plantilla_excel(tmp_path) -> None:
    path = create_inmuebles_excel_template(str(tmp_path / "plantilla.xlsx"))
    workbook = load_workbook(path)
    assert path.endswith("plantilla.xlsx")
    assert workbook.sheetnames == ["Datos", "Ayuda"]


def test_plantilla_tiene_hoja_datos_y_encabezados(tmp_path) -> None:
    path = create_inmuebles_excel_template(str(tmp_path / "plantilla.xlsx"))
    sheet = load_workbook(path)["Datos"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == list(EXPECTED_HEADERS)
    assert sheet.freeze_panes == "A2"


def test_plantilla_tiene_hoja_ayuda_con_codigo_obligatorio(tmp_path) -> None:
    path = create_inmuebles_excel_template(str(tmp_path / "plantilla.xlsx"))
    sheet = load_workbook(path)["Ayuda"]
    values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]
    assert "codigo es obligatorio." in values
    assert any(value == "codigo" for value in values)
    assert any("Obligatoria" == value for value in values)


def test_plantilla_incluye_ejemplos_pedidos(tmp_path) -> None:
    path = create_inmuebles_excel_template(str(tmp_path / "plantilla.xlsx"))
    sheet = load_workbook(path)["Datos"]
    headers = [cell.value for cell in sheet[1]]
    rows = [
        dict(zip(headers, (cell.value or "" for cell in row)))
        for row in sheet.iter_rows(min_row=2, max_row=4)
    ]

    assert rows[0]["codigo"] == "IMP-001"
    assert rows[0]["m2"] == "500"
    assert rows[0]["folio_real"] == ""

    assert rows[1]["codigo"] == "IMP-002"
    assert rows[1]["parcela"] == "P-10"
    assert rows[1]["m2"] == "1234.56"
    assert rows[1]["partida"] == "P-001"
    assert rows[1]["superficie_titulo"] == "1200.50"
    assert rows[1]["fecha_desde"] == "2026-01-01"

    assert rows[2]["codigo"] == "IMP-003"
    assert rows[2]["m2"] == "750"
    assert rows[2]["partida"] == ""


def test_plantilla_documenta_aliases_y_restricciones_de_alcance(tmp_path) -> None:
    path = create_inmuebles_excel_template(str(tmp_path / "plantilla.xlsx"))
    sheet = load_workbook(path)["Ayuda"]
    text = "\n".join(str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value)
    assert "codigo, código, codigo_lote, cod_lote, lote_codigo" in text
    assert "desarrollo, loteo, emprendimiento, barrio" in text
    assert "superficie_titulo" in text
    assert "fecha_desde" in text
    assert "observaciones_catastrales" in text
    assert "No se importan ventas." in text
    assert "No se importa geometría/plano." in text


def test_plantilla_incluye_columnas_avanzadas(tmp_path) -> None:
    path = create_inmuebles_excel_template(str(tmp_path / "plantilla.xlsx"))
    sheet = load_workbook(path)["Datos"]
    headers = [cell.value for cell in sheet[1]]
    for header in (
        "folio_real",
        "circunscripcion",
        "chacra",
        "quinta",
        "fraccion",
        "subparcela",
        "superficie_titulo",
        "superficie_mensura",
        "organismo_origen",
        "fecha_desde",
        "fecha_hasta",
        "estado_dato",
        "observaciones_catastrales",
    ):
        assert header in headers
